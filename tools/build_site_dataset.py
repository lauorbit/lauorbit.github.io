#!/usr/bin/env python3
"""Build a browser-ready ORBIT dataset for the updated static site."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Iterable, Iterator

from openpyxl import load_workbook

JOURNALS_WORKSHEET_NAME = "Unified_Journals"
PUBLISHERS_WORKSHEET_NAME = "Publishers"
CONFERENCES_WORKSHEET_NAME = "Conferences"
ASJC_WORKSHEET_NAME = "ASJC Classification Codes"
RATING_COLUMNS = [
    ("ABDC", "ABDC: 2025 rating"),
    ("JUFO", "JUFO: Level"),
    ("AJG", "AJG: Grade (Converted)"),
    ("FNEGE", "FNEGE: FNEGE_2025"),
    ("VHB", "VHB: Grade (Converted)"),
    ("Norwegian", "Norwegian Level"),
    ("KI", "KI: Level"),
]
RANK_ORDERS = {
    "ABDC": ["C", "B", "A", "A*"],
    "AJG": ["D", "C", "B", "A", "A+"],
    "FNEGE": ["D", "C", "B", "A", "A+"],
    "JUFO": ["0", "1", "2", "3"],
    "KI": ["0", "1", "2", "3"],
    "Norwegian": ["0", "1", "2"],
    "ORBIT": ["D", "C", "B", "A", "A+"],
    "VHB": ["D", "C", "B", "A", "A+"],
}
NUMERIC_RANK_SYSTEMS = {"JUFO", "Norwegian", "KI"}
JUFO_ZERO_ALIASES = {
    "OTHERIDENTIFIEDPUBLICATIONCHANNELS",
    "OTHERIDENTIFIEDPUBLICATIONCHANNEL",
}

FLAG_BUSINESS = 1
FLAG_ELITE = 2
FLAG_WARNING = 4
FLAG_OPEN_ACCESS = 8

RANKING_DENSITY_POINTS = 101
RANKING_DENSITY_MIN_BANDWIDTH = 3.5
RANKING_DENSITY_MAX_BANDWIDTH = 10.0
RANKING_CHART_SYSTEMS = [
    {
        "key": "ABDC",
        "label": "ABDC",
        "filename": "ABDC_JQL.xlsx",
        "column": "2025 rating",
        "color": "#3182CE",
        "description": "ABDC 2025 journal ratings.",
    },
    {
        "key": "AJG",
        "label": "AJG",
        "filename": "AJG_latest_year.xlsx",
        "column": "Grade (Converted)",
        "color": "#DD6B20",
        "description": "AJG latest-year converted grades.",
    },
    {
        "key": "FNEGE",
        "label": "FNEGE",
        "filename": "FNEGE_Journal_Quality_List.xlsx",
        "column": "FNEGE_2025",
        "color": "#D53F8C",
        "description": "FNEGE 2025 grades recognized by the ORBIT pipeline.",
    },
    {
        "key": "JUFO",
        "label": "JUFO",
        "filename": "JUFO_active_export.csv",
        "column": "Level",
        "color": "#E53E3E",
        "description": "Current-year JUFO series entries.",
    },
    {
        "key": "Norwegian",
        "label": "Norwegian Register",
        "filename": "Norwegian_Register.csv",
        "column": "Level",
        "color": "#38A169",
        "description": "Current-year Norwegian journal entries excluding conference proceedings and series.",
    },
    {
        "key": "KI",
        "label": "KI-JL",
        "filename": "KI_Journal_List_2026.xlsx",
        "column": "Level",
        "color": "#0F766E",
        "description": "Karolinska Institutet Journal List (KI-JL) 2026 levels.",
    },
    {
        "key": "VHB",
        "label": "VHB",
        "filename": "VHB_Rating_2024_Area_rating_INT.xlsx",
        "column": "Grade (Converted)",
        "color": "#805AD5",
        "description": "VHB 2024 converted grades.",
    },
]

SCRIPT_DIR = Path(__file__).resolve().parent
SITE_DIR = SCRIPT_DIR.parent
WORKSPACE_ROOT = SITE_DIR.parent

DEFAULT_INPUT = SITE_DIR / "ORBIT.xlsx"
DEFAULT_DATABASES_DIR = (
    WORKSPACE_ROOT
    / "Scripts"
    / "ORBITResult_20260422_092001"
    / "databases"
)
DEFAULT_SCOPUS = WORKSPACE_ROOT / "Scripts" / "ScopusListFeb2026.xlsx"
DEFAULT_META_OUTPUT = SITE_DIR / "data" / "orbit-site-meta.js"
DEFAULT_RECORDS_OUTPUT = SITE_DIR / "data" / "orbit-site-records.js"


def clean_string(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def split_unique(value: object, pattern: str) -> list[str]:
    text = clean_string(value)
    if not text:
        return []

    items: list[str] = []
    seen: set[str] = set()
    for part in re.split(pattern, text):
        token = re.sub(r"\s+", " ", part).strip(" ;|")
        if not token:
            continue
        key = token.casefold()
        if key in seen:
            continue
        seen.add(key)
        items.append(token)
    return items


def split_issns(value: object) -> list[str]:
    tokens: list[str] = []
    seen: set[str] = set()
    for part in split_unique(value, r"[|,;/]+"):
        token = re.sub(r"[^0-9Xx]", "", part).upper()
        if len(token) < 8 or token in seen:
            continue
        seen.add(token)
        tokens.append(token)
    return tokens


def split_identifiers(value: object) -> list[str]:
    tokens: list[str] = []
    seen: set[str] = set()
    for part in split_unique(value, r"[|,;/]+"):
        token = re.sub(r"[^0-9Xx]", "", part).upper()
        if len(token) < 4 or token in seen:
            continue
        seen.add(token)
        tokens.append(token)
    return tokens


def split_urls(value: object) -> list[str]:
    urls: list[str] = []
    seen: set[str] = set()
    for part in split_unique(value, r"\|"):
        if not re.match(r"^https?://", part, flags=re.IGNORECASE):
            continue
        key = part.casefold()
        if key in seen:
            continue
        seen.add(key)
        urls.append(part)
    return urls


def split_asjc_codes(value: object) -> list[str]:
    codes: list[str] = []
    seen: set[str] = set()
    for part in split_unique(value, r"[;|,]+"):
        code = re.sub(r"\D", "", part)
        if len(code) != 4 or code in seen:
            continue
        seen.add(code)
        codes.append(code)
    return codes


def split_rank_tokens(value: object) -> list[str]:
    text = clean_string(value)
    if not text:
        return []

    tokens: list[str] = []
    for part in re.split(r"[|,;/]+", text):
        token = re.sub(r"\s+", "", part).upper()
        if token:
            tokens.append(token)
    return tokens


def is_truthy(value: object) -> bool:
    return clean_string(value).casefold() in {"yes", "true", "1", "y"}


def maybe_float(value: object) -> float | str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        number = float(value)
        return round(number, 4)

    text = clean_string(value)
    if not text:
        return ""
    try:
        return round(float(text), 4)
    except ValueError:
        return text


def sorted_counter(counter: Counter[str]) -> dict[str, int]:
    return {key: counter[key] for key in sorted(counter)}


def normalize_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        number = int(value)
        return f"{number:04d}" if 0 <= number <= 9999 else ""

    digits = re.sub(r"\D", "", str(value))
    if len(digits) != 4:
        return ""
    return digits


def normalize_rank(value: object, system: str) -> str | None:
    tokens = split_rank_tokens(value)
    if not tokens:
        return None

    order = RANK_ORDERS[system]
    if system == "JUFO":
        tokens = ["0" if token in JUFO_ZERO_ALIASES else token for token in tokens]

    if system in NUMERIC_RANK_SYSTEMS:
        numeric_tokens: list[str] = []
        for token in tokens:
            try:
                number = float(token)
            except ValueError:
                continue
            if not math.isfinite(number):
                continue
            numeric_tokens.append(str(int(number)) if number.is_integer() else str(number))
        tokens = numeric_tokens

    valid = [token for token in tokens if token in order]
    if not valid:
        return None

    return max(valid, key=lambda token: order.index(token))


def load_asjc_lookup(scopus_path: Path) -> dict[str, str]:
    workbook = load_workbook(scopus_path, read_only=True, data_only=True)
    worksheet = workbook[ASJC_WORKSHEET_NAME]
    lookup: dict[str, str] = {}

    for code_value, description, *_ in worksheet.iter_rows(values_only=True):
        code = normalize_code(code_value)
        label = clean_string(description)
        if not code or not label:
            continue
        lookup[code] = label

    workbook.close()
    return lookup


def iter_rows(values: Iterable[tuple[object, ...]]) -> Iterable[tuple[object, ...]]:
    for row in values:
        if not row:
            continue
        if all(value is None for value in row):
            continue
        yield row


def iter_excel_records(path: Path, worksheet_name: str | None = None) -> Iterator[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[worksheet_name] if worksheet_name else workbook[workbook.sheetnames[0]]
        headers = [
            clean_string(cell)
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        for raw in iter_rows(worksheet.iter_rows(min_row=2, values_only=True)):
            record: dict[str, object] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                record[header] = raw[index] if index < len(raw) else None
            yield record
    finally:
        workbook.close()


def iter_csv_records(path: Path, delimiter: str) -> Iterator[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        for row in reader:
            yield {
                clean_string(key): clean_string(value)
                for key, value in row.items()
                if key is not None
            }


def percentile(values: list[float], fraction: float) -> float:
    if not values:
        return 0.0
    if len(values) == 1:
        return values[0]

    clamped = min(max(fraction, 0.0), 1.0)
    position = clamped * (len(values) - 1)
    lower = int(math.floor(position))
    upper = int(math.ceil(position))
    if lower == upper:
        return values[lower]

    weight = position - lower
    return values[lower] + ((values[upper] - values[lower]) * weight)


def compute_percentile_midpoints(
    rank_order: list[str],
    counts: Counter[str],
) -> dict[str, float]:
    total = sum(counts.values())
    lower = 0.0
    score_map: dict[str, float] = {}
    for rank in rank_order:
        share = (counts[rank] / total) * 100.0 if total else 0.0
        score_map[rank] = lower + (share / 2.0)
        lower += share
    return score_map


def estimate_density_bandwidth(score_map: dict[str, float], counts: Counter[str]) -> float:
    samples: list[float] = []
    for rank, score in score_map.items():
        samples.extend([score] * counts[rank])

    if len(samples) < 2:
        return 5.0

    samples.sort()
    mean = sum(samples) / len(samples)
    variance = sum((sample - mean) ** 2 for sample in samples) / (len(samples) - 1)
    standard_deviation = math.sqrt(max(variance, 0.0))
    interquartile_range = percentile(samples, 0.75) - percentile(samples, 0.25)
    robust_scale = interquartile_range / 1.34 if interquartile_range > 0 else 0.0
    scale_candidates = [value for value in (standard_deviation, robust_scale) if value > 0]
    if not scale_candidates:
        return 5.0

    scale = min(scale_candidates)
    bandwidth = 0.9 * scale * (len(samples) ** (-0.2))
    return max(
        RANKING_DENSITY_MIN_BANDWIDTH,
        min(RANKING_DENSITY_MAX_BANDWIDTH, bandwidth),
    )


def build_density_curve(
    score_map: dict[str, float],
    counts: Counter[str],
) -> tuple[list[float], list[float], float]:
    x_values = [
        round(index * 100.0 / (RANKING_DENSITY_POINTS - 1), 2)
        for index in range(RANKING_DENSITY_POINTS)
    ]
    total = sum(counts.values())
    if total == 0:
        return x_values, [0.0] * len(x_values), 0.0

    bandwidth = estimate_density_bandwidth(score_map, counts)
    coefficient = 1.0 / (total * bandwidth * math.sqrt(2.0 * math.pi))
    curve: list[float] = []
    for x in x_values:
        density = 0.0
        for rank, count in counts.items():
            if count <= 0:
                continue
            distance = (x - score_map[rank]) / bandwidth
            density += count * math.exp(-0.5 * (distance ** 2))
        curve.append(round(coefficient * density, 6))

    return x_values, curve, round(bandwidth, 4)


def extract_jufo_current_year_level(record: dict[str, str], current_year: int) -> str:
    history = clean_string(record.get("Jufo_History", ""))
    if history:
        for match in re.finditer(r"(\d{4})\s*:\s*([^;]*)", history):
            if int(match.group(1)) == current_year:
                return clean_string(match.group(2))
        return ""

    return clean_string(record.get("Level", ""))


def build_ranking_distributions(
    databases_dir: Path,
    current_year: int,
    orbit_grade_counts: Counter[str],
    fallback_rank_counts: dict[str, Counter[str]],
) -> dict[str, object]:
    systems_payload: list[dict[str, object]] = []
    shared_x_values: list[float] | None = None

    for config in RANKING_CHART_SYSTEMS:
        system = config["key"]
        source_path = databases_dir / config["filename"]
        rank_counts: Counter[str] = Counter()

        if not source_path.exists():
            rank_counts.update(fallback_rank_counts.get(system, Counter()))
        elif system == "JUFO":
            for record in iter_csv_records(source_path, delimiter=","):
                if clean_string(record.get("Type_en")) != "Series":
                    continue
                token = normalize_rank(
                    extract_jufo_current_year_level(record, current_year),
                    system,
                )
                if token is None:
                    continue
                rank_counts[token] += 1
        elif system == "Norwegian":
            year_column = f"Level {current_year}"
            year_column_seen = False
            for record in iter_csv_records(source_path, delimiter=";"):
                if year_column in record:
                    year_column_seen = True
                if clean_string(record.get("Conference Proceedings")) not in {"", "0"}:
                    continue
                if clean_string(record.get("Series")) not in {"", "0"}:
                    continue
                token = normalize_rank(record.get(year_column, ""), system)
                if token is None:
                    continue
                rank_counts[token] += 1

            if not year_column_seen:
                raise ValueError(
                    f"Missing required Norwegian ranking column '{year_column}' in {source_path}"
                )
        else:
            for record in iter_excel_records(source_path):
                token = normalize_rank(record.get(config["column"]), system)
                if token is None:
                    continue
                rank_counts[token] += 1

        rank_order = RANK_ORDERS[system]
        score_map = compute_percentile_midpoints(rank_order, rank_counts)
        x_values, density_curve, bandwidth = build_density_curve(score_map, rank_counts)
        if shared_x_values is None:
            shared_x_values = x_values
        systems_payload.append(
            {
                "key": system,
                "label": config["label"],
                "color": config["color"],
                "description": config["description"],
                "sampleSize": sum(rank_counts.values()),
                "rankCounts": [
                    {"rank": rank, "count": rank_counts[rank]}
                    for rank in rank_order
                    if rank_counts[rank] > 0
                ],
                "scoreMap": {
                    rank: round(score_map[rank], 4)
                    for rank in rank_order
                    if rank_counts[rank] > 0
                },
                "bandwidth": bandwidth,
                "curve": density_curve,
            }
        )

    orbit_counts = Counter(
        {
            "D": orbit_grade_counts.get("D", 0),
            "C": orbit_grade_counts.get("C", 0),
            "B": orbit_grade_counts.get("B", 0),
            "A": orbit_grade_counts.get("A", 0),
            "A+": orbit_grade_counts.get("A+", 0),
        }
    )
    if sum(orbit_counts.values()) > 0:
        systems_payload.append(
            build_distribution_entry(
                key="ORBIT",
                label="ORBIT",
                color="#1F2937",
                description="Final ORBIT grade distribution from the current workbook.",
                rank_counts=orbit_counts,
            )
        )

    return {
        "currentYear": current_year,
        "xValues": shared_x_values or [],
        "systems": systems_payload,
        "method": (
            "Curves are smoothed from current-year tier counts using the same "
            "percentile-midpoint normalization that feeds the ORBIT scoring pipeline."
        ),
    }


def build_distribution_entry(
    *,
    key: str,
    label: str,
    color: str,
    description: str,
    rank_counts: Counter[str],
) -> dict[str, object]:
    rank_order = RANK_ORDERS[key]
    score_map = compute_percentile_midpoints(rank_order, rank_counts)
    _, density_curve, bandwidth = build_density_curve(score_map, rank_counts)

    return {
        "key": key,
        "label": label,
        "color": color,
        "description": description,
        "sampleSize": sum(rank_counts.values()),
        "rankCounts": [
            {"rank": rank, "count": rank_counts[rank]}
            for rank in rank_order
            if rank_counts[rank] > 0
        ],
        "scoreMap": {
            rank: round(score_map[rank], 4)
            for rank in rank_order
            if rank_counts[rank] > 0
        },
        "bandwidth": bandwidth,
        "curve": density_curve,
    }


def build_payload(
    input_path: Path,
    scopus_path: Path,
    databases_dir: Path,
    current_year: int,
) -> tuple[dict[str, object], dict[str, list[list[object]]]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[JOURNALS_WORKSHEET_NAME]
        headers = [
            clean_string(cell)
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        header_index = {header: position for position, header in enumerate(headers)}

        required_columns = {
            "Journal Title",
            "ISSN",
            "URL",
            "Publisher",
            "All Science Journal Classification Codes (ASJC)",
            "Open Access Status",
            "Business Journal",
            "Is Elite Journal",
            "Elite Journal Lists",
            "Is Warning Journal",
            "Warning Journal Lists",
            "KI Weight Eligible",
            "ORBIT_Base_Grade",
            "ORBIT_Grade",
            "Uncertainty Score",
            *[column for _, column in RATING_COLUMNS],
        }
        missing = [column for column in sorted(required_columns) if column not in header_index]
        if missing:
            raise ValueError(f"Missing required columns in {input_path}: {', '.join(missing)}")

        asjc_lookup = load_asjc_lookup(scopus_path)
        rows: list[list[object]] = []
        grade_counts: Counter[str] = Counter()
        source_rank_counts: dict[str, Counter[str]] = {
            label: Counter()
            for label, _ in RATING_COLUMNS
        }
        asjc_counts: Counter[str] = Counter()
        publisher_options: set[str] = set()

        for raw in iter_rows(worksheet.iter_rows(min_row=2, values_only=True)):
            title = clean_string(raw[header_index["Journal Title"]])
            if not title:
                continue

            issns = split_issns(raw[header_index["ISSN"]])
            urls = split_urls(raw[header_index["URL"]])
            publisher = clean_string(raw[header_index["Publisher"]])
            publishers = split_unique(publisher, r"\|")
            asjc_codes = split_asjc_codes(
                raw[header_index["All Science Journal Classification Codes (ASJC)"]]
            )
            open_access_label = clean_string(raw[header_index["Open Access Status"]])
            business = is_truthy(raw[header_index["Business Journal"]])
            elite = is_truthy(raw[header_index["Is Elite Journal"]])
            warning = is_truthy(raw[header_index["Is Warning Journal"]])
            elite_lists = split_unique(raw[header_index["Elite Journal Lists"]], r"\|")
            warning_lists = split_unique(raw[header_index["Warning Journal Lists"]], r"\|")
            ki_weight_eligible = is_truthy(raw[header_index["KI Weight Eligible"]])
            base_grade = clean_string(raw[header_index["ORBIT_Base_Grade"]])
            grade = clean_string(raw[header_index["ORBIT_Grade"]])
            uncertainty = maybe_float(raw[header_index["Uncertainty Score"]])
            title_variants = split_unique(title, r"\|")
            ratings_map = {
                label: clean_string(raw[header_index[column]])
                for label, column in RATING_COLUMNS
            }
            for label, _ in RATING_COLUMNS:
                token = normalize_rank(ratings_map[label], label)
                if token is not None:
                    source_rank_counts[label][token] += 1
            ratings = [ratings_map[label] for label, _ in RATING_COLUMNS]

            flags = 0
            if business:
                flags |= FLAG_BUSINESS
            if elite:
                flags |= FLAG_ELITE
            if warning:
                flags |= FLAG_WARNING
            if open_access_label:
                flags |= FLAG_OPEN_ACCESS

            for code in asjc_codes:
                asjc_counts[code] += 1
            for publisher_token in publishers:
                publisher_options.add(publisher_token)

            grade_counts[grade or "Unranked"] += 1

            rows.append(
                [
                    title,
                    issns,
                    urls,
                    publisher,
                    asjc_codes,
                    flags,
                    ratings,
                    open_access_label,
                    grade,
                    uncertainty,
                    base_grade,
                    ki_weight_eligible,
                    elite_lists,
                    warning_lists,
                ]
            )

        publisher_rows = build_publisher_rows(workbook)
        conference_rows = build_conference_rows(workbook)
    finally:
        workbook.close()

    ranking_distributions = build_ranking_distributions(
        databases_dir,
        current_year,
        grade_counts,
        source_rank_counts,
    )
    client_ranking_distributions = {
        "xValues": ranking_distributions["xValues"],
        "systems": [
            {
                "key": system["key"],
                "label": system["label"],
                "color": system["color"],
                "sampleSize": system["sampleSize"],
                "rankCounts": system["rankCounts"],
                "curve": system["curve"],
            }
            for system in ranking_distributions["systems"]
        ],
    }

    return {
        "flagBits": {
            "business": FLAG_BUSINESS,
            "elite": FLAG_ELITE,
            "warning": FLAG_WARNING,
            "openAccess": FLAG_OPEN_ACCESS,
        },
        "ratingLabels": [label for label, _ in RATING_COLUMNS],
        "publisherOptions": sorted(publisher_options),
        "stats": {
            "gradedEntries": sum(
                count for grade, count in grade_counts.items() if grade != "Unranked"
            ),
            "unrankedEntries": grade_counts.get("Unranked", 0),
            "gradeCounts": sorted_counter(grade_counts),
            "asjcCount": len(asjc_counts),
            "journalEntries": len(rows),
            "conferenceEntries": len(conference_rows),
            "publisherEntries": len(publisher_rows),
            "gradedConferenceEntries": sum(1 for row in conference_rows if row[2] != "Unranked"),
            "gradedPublisherEntries": sum(1 for row in publisher_rows if row[7] != "Unranked"),
        },
        "asjcLookup": {
            code: asjc_lookup[code]
            for code in sorted(asjc_lookup)
            if code in asjc_counts
        },
        "rankingDistributions": client_ranking_distributions,
    }, {
        "journals": rows,
        "publishers": publisher_rows,
        "conferences": conference_rows,
    }


def build_publisher_rows(workbook) -> list[list[object]]:
    worksheet = workbook[PUBLISHERS_WORKSHEET_NAME]
    headers = [
        clean_string(cell)
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    header_index = {header: position for position, header in enumerate(headers)}
    required_columns = {
        "Final_Grade",
        "Reliability_Score",
        "JUFO_ISBNs",
        "NOR_ISBNs",
        "JUFO_JUFOLevel",
        "NOR_NorwegianLevel",
        "JUFO_Name",
        "NOR_InternationalTitle",
        "NOR_URL",
        "Shared_ISBNs",
    }
    missing = [column for column in sorted(required_columns) if column not in header_index]
    if missing:
        raise ValueError(
            f"Missing required columns in {PUBLISHERS_WORKSHEET_NAME}: {', '.join(missing)}"
        )

    rows: list[list[object]] = []
    for raw in iter_rows(worksheet.iter_rows(min_row=2, values_only=True)):
        jufo_name = clean_string(raw[header_index["JUFO_Name"]])
        nor_title = clean_string(raw[header_index["NOR_InternationalTitle"]])
        aliases = split_unique("|".join(value for value in [nor_title, jufo_name] if value), r"\|")
        display_name = aliases[0] if aliases else ""
        jufo_isbns = split_identifiers(raw[header_index["JUFO_ISBNs"]])
        nor_isbns = split_identifiers(raw[header_index["NOR_ISBNs"]])
        shared_isbns = split_identifiers(raw[header_index["Shared_ISBNs"]])
        if not display_name and not jufo_isbns and not nor_isbns and not shared_isbns:
            continue

        rows.append(
            [
                display_name or "Untitled publisher",
                aliases,
                jufo_isbns,
                nor_isbns,
                shared_isbns,
                clean_string(raw[header_index["JUFO_JUFOLevel"]]),
                clean_string(raw[header_index["NOR_NorwegianLevel"]]),
                clean_string(raw[header_index["Final_Grade"]]) or "Unranked",
                maybe_float(raw[header_index["Reliability_Score"]]),
                clean_string(raw[header_index["NOR_URL"]]),
            ]
        )

    return rows


def build_conference_rows(workbook) -> list[list[object]]:
    worksheet = workbook[CONFERENCES_WORKSHEET_NAME]
    headers = [
        clean_string(cell)
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    header_index = {header: position for position, header in enumerate(headers)}
    required_columns = {
        "canonical_issn",
        "name of the conference",
        "ORBIT Grade",
        "uncertainty score",
        "CORE rank",
        "JUFO Level",
        "Norwegian level",
    }
    missing = [column for column in sorted(required_columns) if column not in header_index]
    if missing:
        raise ValueError(
            f"Missing required columns in {CONFERENCES_WORKSHEET_NAME}: {', '.join(missing)}"
        )

    rows: list[list[object]] = []
    for raw in iter_rows(worksheet.iter_rows(min_row=2, values_only=True)):
        name = clean_string(raw[header_index["name of the conference"]])
        if not name:
            continue

        rows.append(
            [
                name,
                split_issns(raw[header_index["canonical_issn"]]),
                clean_string(raw[header_index["ORBIT Grade"]]) or "Unranked",
                maybe_float(raw[header_index["uncertainty score"]]),
                clean_string(raw[header_index["CORE rank"]]),
                clean_string(raw[header_index["JUFO Level"]]),
                clean_string(raw[header_index["Norwegian level"]]),
            ]
        )

    return rows


def to_site_url(path: Path) -> str:
    relative_path = path.resolve().relative_to(SITE_DIR.resolve())
    return f"./{relative_path.as_posix()}"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--databases-dir", type=Path, default=DEFAULT_DATABASES_DIR)
    parser.add_argument("--scopus", type=Path, default=DEFAULT_SCOPUS)
    parser.add_argument("--meta-output", type=Path, default=DEFAULT_META_OUTPUT)
    parser.add_argument("--records-output", type=Path, default=DEFAULT_RECORDS_OUTPUT)
    args = parser.parse_args()

    meta_payload, rows_payload = build_payload(
        input_path=args.input.resolve(),
        scopus_path=args.scopus.resolve(),
        databases_dir=args.databases_dir.resolve(),
        current_year=datetime.now().year,
    )
    meta_output = args.meta_output.resolve()
    records_output = args.records_output.resolve()
    meta_output.parent.mkdir(parents=True, exist_ok=True)
    records_output.parent.mkdir(parents=True, exist_ok=True)

    meta_payload["recordsScriptUrl"] = to_site_url(records_output)

    serialized_meta = json.dumps(meta_payload, ensure_ascii=True, separators=(",", ":"))
    meta_output.write_text(
        "// Auto-generated by tools/build_site_dataset.py\n"
        f"globalThis.ORBIT_SITE_META={serialized_meta};\n",
        encoding="utf-8",
    )
    serialized_rows = json.dumps(rows_payload, ensure_ascii=True, separators=(",", ":"))
    records_output.write_text(
        "// Auto-generated by tools/build_site_dataset.py\n"
        f"globalThis.ORBIT_SITE_ROWS={serialized_rows};\n",
        encoding="utf-8",
    )

    print(
        "Wrote site dataset with "
        f"{len(rows_payload['journals'])} journals, "
        f"{len(rows_payload['conferences'])} conferences, and "
        f"{len(rows_payload['publishers'])} publishers to {meta_output} and {records_output}"
    )


if __name__ == "__main__":
    main()
